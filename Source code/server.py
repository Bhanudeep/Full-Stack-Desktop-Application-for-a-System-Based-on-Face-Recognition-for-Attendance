import datetime,face_recognition
from flask import Flask, flash, request, redirect, url_for,render_template,session
from flask_pymongo import PyMongo
import os,shutil
import bcrypt
import pandas as pd
import openpyxl
from PIL import Image
import PIL
import face_recognition.api as face_recognition
import numpy 
from datetime import datetime,date
import smtplib
import random
import math
from resizeimage import resizeimage
import json
import pandas as pd
from PIL import Image
import socket
from tkinter import *
from tkinter.filedialog import askopenfilenames
s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
s.connect(("8.8.8.8", 80))
ip_addr=s.getsockname()[0]
s.close()
app = Flask(__name__, static_url_path='/static')
app.secret_key = '?&Kjhd$^ljm>x21'
app.config['MONGO_DBNAME'] = 'FacultyLogin'
app.config['MONGO_URI'] = 'mongodb://127.0.0.1:27017/FacultyLogin'
mongo = PyMongo(app)  
day=int(date.today().day)
mnth=str(date.today().month)
if(int(mnth)<10):
    mnth=str('0'+mnth)
yr=str(date.today().year)
today = str(date.today())
bad_chars = [';', ':', '!', "*", " ","'","]","["]
def view_attendance1(cur_sheetpath,xlhtmlpath):
	wb=pd.read_excel(cur_sheetpath)
	wb.to_html(xlhtmlpath)

def view_absentees(cur_sheetpath,h):
	l=[]
	l.clear()
	wb=openpyxl.load_workbook(cur_sheetpath)
	ws=wb.active
	m_row=ws.max_row
	for i in range(2,m_row+1):
		if ws.cell(column=h,row=i).value == 0:
			l.append(ws.cell(column=1,row=i).value)
	return l


def compute(ws,day,h,c,l,wb,excelpath,xlhtmlpath,time,subcode):
    now = datetime.now()
    ws=wb.active
    m_row = ws.max_row 
    ws.cell(row = 1, column = c).value = str(subcode)+"\n"+str(h)+"\n"+time
    ws.column_dimensions['A'].width = 15
    ws.row_dimensions[1].height = 38
    sum=0
    for i in range(2, m_row + 1):
        for j in l:
            for p in bad_chars :
                j=str(j)
                j = j.replace(p, '')
        for j in l:
            if(str(ws.cell(row = i, column = 1).value) == j):
                ws.cell(row = i, column = c).value = 1
        if(ws.cell(row = i, column = c).value == None):
            ws.cell(row = i, column = c).value = 0
        sum+=int(ws.cell(row = i, column = c).value)
    if(day > 0 and day < 16):
        wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        view_attendance1(cur_sheetpath,xlhtmlpath)
        l=view_absentees(cur_sheetpath,c) # c=h+1
    else:
        wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
        view_attendance1(cur_sheetpath,xlhtmlpath)
        l=view_absentees(cur_sheetpath,c) # c=h+1
    return sum,l

def create_sheet(l,h,excelpath,xlhtmlpath,time,subcode):    
    h=float(h)
    h=int(h)
    if(day == 1 or day == 16):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = today
        wb_temp = openpyxl.load_workbook(excelpath+"\\template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        m_row = ws_temp.max_row
        for i in range(1, m_row + 1):
            for j in range(1,m_column+1):
                ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
        wb.save(excelpath+today+".xlsx")
        sum,l=compute(ws,day,h,h+m_column,l,wb,excelpath,xlhtmlpath,time,subcode)
     
    elif(day > 1 and day < 16):
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        wb_temp = openpyxl.load_workbook(excelpath+"\\template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        sum,l=compute(ws,day,h,h+m_column,l,wb,excelpath,xlhtmlpath,time,subcode)
        
        
    else:
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        wb_temp = openpyxl.load_workbook(excelpath+"\\template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames 
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        sum,l=compute(ws,day,h,h+m_column,l,wb,excelpath,xlhtmlpath,time,subcode)
    return sum,l

def match(opath,epath,erpath,departments,year,sub,hour,subcode):
    absent=[]
    present=0
    depp=departments[0]
    if depp=="NONE":
        try:
            now = datetime.now()
            time=str(now.strftime("%I:%M %p"))
            project=str(os.getcwd()+"\\")
            default=str(project+"static\\images\\")
            originallist={}
            for department in departments:
                jsonpath=str(default+"Subject"+"\\"+sub+"\\original.json")
                data = open(jsonpath,"r")
                n=data.read()
                temp_list=json.loads(n)
                originallist.update(temp_list)
            for i in originallist:
                originallist[i]=numpy.asarray(originallist[i])
            l=[]
            for a in range(0,2):
                #extracted images encoding
                extractedlist=[]
                for filename in os.listdir(epath):
                    sublist=[]
                    filename=str(filename)
                    sublist.append(filename)
                    file_path=epath+filename
                    image=face_recognition.load_image_file(file_path)
                    image_encoding=face_recognition.face_encodings(image,num_jitters=5)
                    if len(image_encoding)>0:
                        image_encoding=image_encoding[0]
                        sublist.append(image_encoding)
                        extractedlist.append(sublist)
                    else:
                        shutil.move(epath+filename,erpath)
                
                for key in originallist:
                    rollno=key[0:10]
                    if rollno in l:
                        continue 
                    for esublist in extractedlist:
                        oimg=originallist[key]
                        eimg=esublist[1]
                        result=face_recognition.compare_faces([oimg],eimg,tolerance=0.435)
                        if result[0]== True:
                            l.append(rollno)
                            os.remove(epath+esublist[0])
                            extractedlist.remove(esublist)
                            break
            absent=[]
            present=0
            for department in departments:
                xl_fname=sub+".html"
                excelpath=str(default+"\\Subject\\"+sub+"\\excel\\")
                xlhtmlpath=str(project+"templates/excel_html/"+xl_fname)
                c,a=create_sheet(l,hour,excelpath,xlhtmlpath,time,subcode) 
                absent+=a
                present+=c
            return present,absent
        except:
            # flash("Please Separate Encodings first. If you are viewing this message you either haven't \n separated Encodings, if they are separated retake photo again.")
            return present,absent
    else:
        try:
            now = datetime.now()
            time=str(now.strftime("%I:%M %p"))
            project=str(os.getcwd()+"\\")
            default=str(project+"static\\images\\")
            originallist={}
            for department in departments:
                jsonpath=str(default+department+"\\"+year+"\\original.json")
                data = open(jsonpath,"r")
                n=data.read()
                temp_list=json.loads(n)
                originallist.update(temp_list)
            for i in originallist:
                originallist[i]=numpy.asarray(originallist[i])
            l=[]
            for a in range(0,2):
                #extracted images encoding
                extractedlist=[]
                for filename in os.listdir(epath):
                    sublist=[]
                    filename=str(filename)
                    sublist.append(filename)
                    file_path=epath+filename
                    image=face_recognition.load_image_file(file_path)
                    image_encoding=face_recognition.face_encodings(image,num_jitters=5)
                    if len(image_encoding)>0:
                        image_encoding=image_encoding[0]
                        sublist.append(image_encoding)
                        extractedlist.append(sublist)
                    else:
                        shutil.move(epath+filename,erpath)
                
                for key in originallist:
                    rollno=key[0:10]
                    if rollno in l:
                        continue 
                    for esublist in extractedlist:
                        oimg=originallist[key]
                        eimg=esublist[1]
                        result=face_recognition.compare_faces([oimg],eimg,tolerance=0.435)
                        if result[0]== True:
                            l.append(rollno)
                            os.remove(epath+esublist[0])
                            extractedlist.remove(esublist)
                            break
            absent=[]
            present=0
            for department in departments:
                xl_fname=department+"_"+year+"_"+sub+".html"
                excelpath=str(default+department+"/"+year+"/subjects/"+sub+"/excel/")
                xlhtmlpath=str(project+"templates/excel_html/"+xl_fname)
                c,a=create_sheet(l,hour,excelpath,xlhtmlpath,time,subcode) 
                absent+=a
                present+=c
            return present,absent
        except:
            flash("Please Separate Encodings first. If you are viewing this message you either haven't \n separated Encodings, if they are separated retake photo again.")
            return present,absent
    

def view_attendance_date(date,excelpath,xlhtmlpath):
	date=str(date)
	yr=date[0:4]
	mnth=date[5:7]
	day=int(date[8:])
	if(day < 16):
		try:
			wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-01.xlsx")
			ws = wb.active
			sheet = str(wb[date])
			d = sheet[12:22]
			re=pd.read_excel(excelpath+yr+"-"+mnth+"-01.xlsx",sheet_name=date)
			re.to_html(xlhtmlpath)
			return True,d,""
		except:
			return False,0,"Given date may not be a working day"
	else:
		try:
			wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-16.xlsx")
			ws = wb.active
			sheet = str(wb[date])
			d = sheet[12:22]
			re=pd.read_excel(excelpath+yr+"-"+mnth+"-16.xlsx",sheet_name=date)
			re.to_html(xlhtmlpath)
			return True,d,""
		except:
			return False,0,"Given date may not be a working day"




def mail(email,fp):
    digits = [i for i in range(0, 10)]
    OTP = ""
    for i in range(6):
        index = math.floor(random.random() * 10)
        OTP += str(digits[index])
    if fp==1:
        content = '\nHello! \nyou recently requested for password change please enter the given OTP to reset your password : '+ OTP +' .\nIf you did not request a password reset, please ignore this email or reply to let us know.This reset is only valid for the next 30 minutes.\n\nThanks....\nFaceRecognition based attendance system NMREC'
        username = "8888888888"
        password = "*********"
        sender = "attendance@nmrec.edu.in"
        recipient = email
        mail = smtplib.SMTP("smtp.gmail.com",587)
        mail.ehlo() 
        mail.starttls() 
        mail.login("attendance@nmrec.edu.in","nmrec@frba")
        header = 'To:' + recipient + '\n' + 'From:' + sender + '\n' + 'Subject: Reset password OTP \n'
        content = header+content
        mail.sendmail(sender,recipient,content)
        mail.close
    return(OTP)

def json_extractor(dep,year,sub,default,path):
	m=""
	if os.path.exists(default+dep+'\\'+year+'\\original.json'):
		org_json = open(default+dep+'\\'+year+'\\original.json','r')
	else:
		m="Face Encodings of all students of "+dep+"-"+year+" year are missing"
		return m 
	data = json.load(org_json)
	if os.path.exists(path+'excel\\template.xlsx'):
		wb = openpyxl.load_workbook(path+'excel\\template.xlsx')
	else:
		m="Excelsheet template of "+dep+"-"+year+" year "+sub+" subject is missing"
		return m
	ws = wb.active
	m_row = ws.max_row
	dictionary={}
	for i in range(2,m_row+1):
		for key in data:
			roll_no = key[0:10]
			if(ws.cell(row = i, column = 1).value == roll_no):
				temp_dictionary = {roll_no : data[key]}
				dictionary.update(temp_dictionary)
				json_path=str(path+"original.json")
				json_file=open(json_path,"w")
				d=json.dumps(dictionary)
				json_file.write(d)
				org_json.close()
	return m 

def Encodings(opath,erpath,jsonpath):
	dict1={}
	f=open(jsonpath,"w")
	error_image_name=[]
	for filename in os.listdir(opath):
		sublist=[]
		filename=str(filename)
		sublist.append(filename)
		imgpath=opath+filename
		image=face_recognition.load_image_file(imgpath)
		image_encoding=face_recognition.face_encodings(image,num_jitters=100)
		if len(image_encoding)<=0:
			img=Image.open(imgpath)
			l=img.size
			m=max(l)
			img=resizeimage.resize_crop(img,[m,m])
			img.save(imgpath,img.format)
			out=Image.open(imgpath)
			out=out.rotate(270)
			out.save(imgpath)
			image = face_recognition.load_image_file(imgpath)
			image_encoding=face_recognition.face_encodings(image,num_jitters=100)
		if len(image_encoding)>0:
			image_encoding=image_encoding[0]
			dict1.update({filename:list(image_encoding)})
		else:
			shutil.move(opath+filename,erpath)
			error_image_name.append(filename[0:10])
	d=json.dumps(dict1)
	f.write(d)
	return error_image_name

def extract_count(gpath,epath,filename):
	image = face_recognition.load_image_file(gpath)
	face_locations = face_recognition.face_locations(image,number_of_times_to_upsample=0)
	count=len(face_locations)
	if count==0:
		img=Image.open(gpath)
		l=img.size
		m=max(l)
		img=resizeimage.resize_crop(img,[m,m])
		img.save(gpath,img.format)
		out=Image.open(gpath)
		out=out.rotate(270)
		out.save(gpath)
		image = face_recognition.load_image_file(gpath)
		face_locations = face_recognition.face_locations(image)
		count=len(face_locations)
	if count>0:
		i=0
		for face_location in face_locations:
			i=i+1
			top, right, bottom, left = face_location
			face_image = image[top:bottom, left:right]
			pil_image = PIL.Image.fromarray(face_image)
			name =filename+ "-"+str(i) +'.jpg'
			pil_image.save(epath+name)  
	return count 

def Enco_lst():
    class my_dictionary(dict): 
    
        # __init__ function 
        def __init__(self): 
            self = dict() 
            
        # Function to add key:value 
        def add(self, key, value): 
            self[key] = value 

        def add_values_in_dict(self, key, list_of_values):
            if key not in self:
                self[key] = list()
            self[key].extend(list_of_values)
            return self

    new_dic=my_dictionary()
    dic_obj=my_dictionary()
    path=os.getcwd()
    rootdir = str(path)+"\\static\\images"
    sub_list=[]
    exis=[]
    for file in os.listdir(rootdir):
        d = os.path.join(file)
        sub_list.append(d)
    for listr in sub_list:
        yeardir = str(path)+"\\static\\images"+"\\"+listr
        yr_list=[]
        for file in os.listdir(yeardir):
            d = os.path.join(file)
            dic_obj.key=listr
            dic_obj.value=d
            dic_obj.add_values_in_dict(dic_obj.key, dic_obj.value)
    for liste in dic_obj:
        r=dic_obj[liste]
        for e in r:
            pathw=rootdir+"\\"+str(liste)+"\\"+e+"\\original.json"
            if os.path.exists(pathw):
                exis.append(e)
                new_dic.key=liste
                new_dic.value=e
                new_dic.add_values_in_dict(new_dic.key, new_dic.value)
    return new_dic   # In this example, foo can return a string, list or dict

@app.route('/', methods=['POST','GET'])
def home():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user['username']=="ADMIN":
            return render_template('/admin/admin.html',username=name)
        return render_template('details.html',username=name)
    return render_template('login.html')

@app.route('/login', methods=['POST','GET'])
def login():
    if 'username' in session:
        return """<center><h2>Please logout to login as another user...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        alert="alert"
        if request.method == 'POST':
            users = mongo.db.users
            login_user = users.find_one({'username' : request.form['UserName']})
            login_mail=users.find_one({'mailid': request.form['UserName']})
            if login_user:
                if bcrypt.hashpw(request.form['Password'].encode('utf-8'), login_user['password']) == login_user['password']:
                    user = login_user['username']
                    session['username'] = user
                    return redirect(url_for('home'))
            elif login_mail:
                if bcrypt.hashpw(request.form['Password'].encode('utf-8'), login_mail['password']) == login_mail['password']:
                    user = login_mail['username']
                    session['username'] = user
                    return redirect(url_for('home'))
            flash("Username or Password is invalid")
        return render_template("login.html",failed=alert)

@app.route('/register', methods=['POST', 'GET'])
def register():
    if 'username' in session:
        return """<center><h2>Please logout to register as a new user...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        alert="alert"
        if request.method == 'POST':
            users = mongo.db.users
            OTPS = mongo.db.OTPS
            Admin = users.find_one({'username' : 'ADMIN'})
            if(Admin):
                Admin_mailid = Admin['mailid']
                passcode = OTPS.find_one({'mailid': Admin_mailid})
            existing_user = users.find_one({'username' : request.form['UserName']})
            existing_mail= users.find_one({'mailid': request.form['Mailid']})
            if request.form['UserName']== "ADMIN" or (passcode and (bcrypt.hashpw(request.form['passcode'].encode('utf-8'), passcode['OTP'])) == passcode['OTP']):
                if existing_user is None and existing_mail is None:
                    hashpass = bcrypt.hashpw(request.form['Password'].encode('utf-8'), bcrypt.gensalt())
                    users.insert_one({'username' : request.form['UserName'], 'password' : hashpass , 'firstname' : request.form['First_Name'] , 'lastname' : request.form['Last_Name'], 'mailid': request.form['Mailid']})
                    flash('successfully registered ...')
                    return render_template("login.html",passed=alert)
                flash('Username or Mailid is already exist')
                return render_template('register.html')
            flash('Invalid Passcode')
        return render_template('register.html')

@app.route("/forgot",methods=['POST','GET'])
def forgot():
    if 'username' in session:
        return """<center><h2>Unautnorised Access found please logout and try again...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        if request.method == 'POST':
            users = mongo.db.users
            OTPS = mongo.db.OTPS
            mailid = request.form['mailid']
            valid_mailid = users.find_one({'mailid' : request.form['mailid']})
            existing_mail = OTPS.find_one({'mailid' : request.form['mailid']})
            if valid_mailid:
                OTPO = mail(mailid,1)
                OTP = bcrypt.hashpw(OTPO.encode('utf-8'), bcrypt.gensalt())
                if existing_mail:
                    OTPS.update_one({'mailid' : request.form['mailid']},{"$set" : {"OTP": OTP}})
                else:
                    OTPS.insert_one({'OTP' : OTP , 'mailid' : request.form['mailid']})
                return render_template("resetpassword.html",mailid = mailid)
            flash('Invalid Mail id')
        return render_template("forgot.html")


@app.route("/resetpassword",methods=['POST','GET'])
def reset():
    if 'username' in session:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        alert="alert"
        if request.method == 'POST':
            users = mongo.db.users
            OTPS = mongo.db.OTPS
            User_mail = OTPS.find_one({'mailid': request.form['mailid']})
            if User_mail:
                if(bcrypt.hashpw(request.form['OTP'].encode('utf-8'), User_mail['OTP'])) == User_mail['OTP']:
                    newpassword = bcrypt.hashpw(request.form['Password'].encode('utf-8'), bcrypt.gensalt())
                    users.update_one({'mailid':request.form['mailid']},{"$set":{"password":newpassword}})
                    OTPS.remove({'mailid':request.form['mailid']},True)
                    flash('Password reset successfull')
                    return render_template("login.html",passed=alert)
                flash('Invalid OTP')
            return render_template("resetpassword.html",mailid = User_mail['mailid'])
        else:
            return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""

@app.route("/logout",methods=["POST","GET"])
def logout():
    session.clear()
    return redirect(url_for("login"))

def createpath(dep,year,subj_code,hr):
    dep=dep.upper()
    year=year.upper()
    subj_code=subj_code.upper()
    temp=dep+"\\"+year+"\\"
    hr=hr.upper()
    opath=str(default+temp+"original\\")
    ojsonpath=str(default+temp+"original.json")
    erpath=str(default+temp+"error\\")
    # Subjects folder Inner folders paths
    gpath=str(default+temp+"subjects\\"+subj_code+"\\group\\"+hr+"\\")
    epath=str(default+temp+"subjects\\"+subj_code+"\\extracted\\")
    extracted_erpath=str(default+temp+"subjects\\"+subj_code+"\\error")
    cjsonpath=str(default+temp+"subjects\\"+subj_code+"\\original.json")
    return opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath


project=os.getcwd()+"\\"             
default=project+"static\\images\\"    
if not os.path.isdir(str(default)):
    os.mkdir(str(default))

@app.route('/path', methods=['GET', 'POST'])
def checkpath():
    alert="alert"
    if request.method == "POST":
        dep=request.form["department"]
        year=request.form["year"]
        sub=request.form["subject"]
        hr=request.form["hour"]
        dep=dep.upper()
        sub=sub.upper()
        departments=dep.split(",")
        subcode=request.form["subcode"]
        for i in range(len(departments)):
            department=departments[i]
            opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath=createpath(department,year,sub,hr)
            flag=0
            if(os.path.isdir(gpath)):
                flag=1
                for filename in os.listdir(epath):
                    ipath=str(epath+filename)
                    os.remove(ipath)
                for filename in os.listdir(extracted_erpath):
                    ipath=str(extracted_erpath+filename)
                    os.remove(ipath)
            else:
                flag=0
        if(flag):
            return render_template('upload.html',dep=dep,year=year,sub=sub,hour=hr,subcode=subcode)
        else:
            flash("Ivalid details")
    
            return render_template('details.html',fail=alert)
    else:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/path_com', methods=['GET', 'POST'])
def checkpathcom():
    alert="alert"
    if request.method == "POST":
        dep="None"
        year="None"
        sub=request.form["subject"]
        hr=request.form["hour"]
        dep=dep.upper()
        sub=sub.upper()
        try:
            return render_template('upload.html',dep=dep,year=year,sub=sub,hour=hr)
        except:
            flash("Ivalid details")
    
            return render_template('details.html',fail=alert)
    else:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/view_attendance',methods=['GET','POST'])
def view_attendance():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if request.method == "POST":
            alert="alert"
            try:
                dep=request.form["department"]
                year=request.form["year"]
                sub=request.form["subject"]
            except:
                sub=request.form["subject"]
            date=request.form["date"]
            dep=dep.upper()
            year=year.upper()
            sub=sub.upper()
            if dep=="NONE":
                success="success"
                check_path=str(default+"\\Subject\\"+sub+"\\")
                if(os.path.isdir(check_path)):
                    xl_fname=sub+".html"
                    xlhtmlpath=str(project+"templates\\excel_html\\"+xl_fname)
                    excelpath=str(default+"\\Subject\\"+sub+"\\excel\\")
                    value,dt,msg= view_attendance_date(date,excelpath,xlhtmlpath )
                    if login_user["username"] == "ADMIN":
                        return render_template("/admin/view_attendance.html",dep=dep,year=year,sub=sub,view_excel=value,date_msg=msg,date=dt,username=name,success=success)    
                    return render_template("details.html",dep=dep,year=year,sub=sub,view_excel=value,date_msg=msg,date=dt,username=name)
                else:
                    flash("Invalid ww details")
                    if login_user["username"] == "ADMIN":
                        return render_template("/admin/view_attendance.html",username=name,fail=alert)
                    return render_template("details.html",username=name,fail=alert)
            else:

                check_path=str(default+dep+"\\"+year+"\\subjects\\"+sub+"\\")
                if(os.path.isdir(check_path)):
                    xl_fname=dep+"_"+year+"_"+sub+".html"
                    xlhtmlpath=str(project+"templates\\excel_html\\"+xl_fname)
                    excelpath=str(default+dep+"\\"+year+"\\subjects\\"+sub+"\\excel\\")
                    value,dt,msg= view_attendance_date(date,excelpath,xlhtmlpath )
                    if login_user["username"] == "ADMIN":
                        return render_template("/admin/view_attendance.html",dep=dep,year=year,sub=sub,view_excel=value,date_msg=msg,date=dt,username=name)    
                    return render_template("details.html",dep=dep,year=year,sub=sub,view_excel=value,date_msg=msg,date=dt,username=name)
                else:
                    flash("Invalid details")
                    if login_user["username"] == "ADMIN":
                        return render_template("/admin/view_attendance.html",username=name,fail=alert)
                    return render_template("details.html",username=name,fail=alert)
        if login_user["username"] == "ADMIN":
            return render_template("/admin/view_attendance.html",username=name)
        return render_template("details.html",username=name)
    return render_template("login.html")


@app.route('/upload',methods=['GET','POST'])
def imgupload():
    success="success"
    if request.method == "POST":
        dep=request.form["department"]
        sub=request.form["subject"]
        hr=request.form["hour"]
        year=request.form["year"]
        total=request.form["total"]
        subcode=request.form["subcode"]
        if dep=="NONE":
            image = request.files["image"]
            filename=str(os.getcwd()+"\\static\\images\\Subject\\"+sub+"\\group\\"+hr+"\\")
            try:
                os.makedirs(filename)
            except:
                print()
            app.config["IMAGE_UPLOADS"]=filename
            filename1="image.jpg"
            image.save(os.path.join(app.config["IMAGE_UPLOADS"], filename1))
            app.config["IMAGE_UPLOADS"]=filename
            flash("Your image is uploaded successfully")
            
            return render_template("upload.html",success=success,dep=dep,year=year,sub=sub,hour=hr,count=1,total=1,subcode=subcode)
        else:
            
            departments=dep.split(",")
            filename=str(date.today())
            if request.files:
                for department in departments:
                    opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath=createpath(department,year,sub,hr)
                    image = request.files["image"]
                    i=1
                    while(i):
                        fpath1=gpath+filename+"_"+str(i)+".png"
                        fpath2=gpath+filename+"_"+str(i)+".jpg"
                        if(os.path.isfile(fpath1) or os.path.isfile(fpath2)):
                            i+=1
                            continue
                        else:
                            app.config["IMAGE_UPLOADS"]=gpath
                            filename=filename+"_"+str(i)
                            filename1=filename+".jpg"
                            image.save(os.path.join(app.config["IMAGE_UPLOADS"], filename1))
                            for department in departments:
                                if department!=departments[0]:
                                    dest_path=str(default+department+"\\"+year+"\\subjects\\"+sub+"\\group\\"+hr+"\\")
                                    shutil.copy(fpath2,dest_path)
                            count=extract_count(fpath2,epath,filename)
                            flash("Your image is uploaded successfully")
                            return render_template("upload.html",success=success,dep=dep,year=year,sub=sub,hour=hr,count=count,total=total,subcode=subcode)
                            break
    else:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""

@app.route('/discard',methods=['GET','POST'])
def discard():
    if request.method == "POST":
        dep=request.form["department"]
        year=request.form["year"]
        sub=request.form["subject"]
        hr=request.form["hour"]
        count=request.form["count"]
        total=request.form["total"]
        count=float(count)
        total=float(total)
        departments=dep.split(",")
        for department in departments:
            opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath=createpath(department,year,sub,hr)
            filename=str(date.today())
            i=1
            while(i):
                fpath1=gpath+filename+"_"+str(i)+".jpg"
                fpath2=gpath+filename+"_"+str(i+1)+".jpg"
                if(os.path.isfile(fpath2)):
                    i+=1
                    continue
                else:
                    os.remove(fpath1)
                    break;
            if(departments.index(department)==0):
                while(int(count)>=1):
                    fpath=epath+filename+"_"+str(i)+"-"+str(int(count))+".jpg"
                    os.remove(fpath)
                    count=int(count-1)
        total=int(total-count)
        return render_template("upload.html",dep=dep,year=year,sub=sub,hour=hr,total=total)
    else:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/mark_attendance',methods=['GET','POST'])
def mark_attendance():
    success="success"
    if request.method == "POST":
        dep=request.form["department"].upper()
        year=request.form["year"].upper()
        sub=request.form["subject"].upper()
        hr=request.form["hour"].upper()
        departments=dep.split(",")
        hr=request.form["hour"]
        subcode=request.form["subcode"].upper()
        if dep=="NONE":
            epath=os.getcwd()+"\\static\\images\\Subject\\"+sub+"\\extracted\\"
            extracted_erpath=os.getcwd()+"\\static\\images\\Subject\\"+sub+"\\error\\"
            opath=os.getcwd()+"\\static\\images\\Subject\\"+sub+"\\extracted\\"
            present,l=match(opath,epath,extracted_erpath,departments,year,sub,hr,subcode)
            images=[]
            for filename in os.listdir(epath):
                ipath=str("\\static\\images\\Subject\\"+sub+"\\extracted\\"+filename)
                images.append(ipath)
            for filename in os.listdir(extracted_erpath):
                ipath=str("\\static\\images\\Subject\\"+sub+"\\error\\"+filename)
                images.append(ipath)
            flash("Period-"+str(hr)+" Attendance of "+sub+" subject is marked successfully for "+str(present)+" students")
            success="success1"
            return render_template("error.html",sub=sub,success=success)
        else:
            opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath=createpath(departments[0],year,sub,hr)
            present,l=match(opath,epath,extracted_erpath,departments,year,sub,hr,subcode)
            present=str(present)
            images=[]
            for filename in os.listdir(epath):
                ipath=str("\\static\\images\\"+dep+"\\"+year+"\\subjects\\"+sub+"\\extracted\\"+filename)
                images.append(ipath)
            for filename in os.listdir(extracted_erpath):
                ipath=str("\\static\\images\\"+dep+"\\"+year+"\\subjects\\"+sub+"\\error\\"+filename)
                images.append(ipath)
            flash("Period-"+hr+" Attendance of "+dep+"-"+year+" " +sub+" subject is marked successfully for "+present+" students")
            return render_template("error.html",dep=dep,departments_list=departments,year=year,sub=sub,hour=hr,present=present,success=success,images=images,absent_list=l)
    else:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/retake',methods=['GET','POST'])
def retake():
    success="success"
    if request.method == "POST":
        dep=request.form["department"]
        year=request.form["year"]
        sub=request.form["subject"]
        hr=request.form["hour"]
        total=request.form["total"]
        departments=dep.split(",")
        locations = mongo.db.locations
        location_value = locations.find({})
        for loc in location_value:
            location_list=loc['locations']
        opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath=createpath(departments[0],year,sub,hr)
        try:
            for filename in os.listdir(epath):
                ipath=str(project+"static\\images\\"+departments[0]+"\\"+year+"\\subjects\\"+sub+"\\extracted\\"+filename)
                os.remove(ipath)
            for filename in os.listdir(erpath):
                ipath=str(project+"static\\images\\"+departments[0]+"\\"+year+"\\subjects\\"+sub+"\\error\\"+filename)
                os.remove(ipath)
        except:
            print()
        return render_template("upload.html",dep=dep,year=year,sub=sub,hour=hr,total=total)
    else:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""

@app.route('/admin_page',methods=['GET','POST'])
def admin_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/admin.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/Hours_page',methods=['GET','POST'])
def Hours_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/hours.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""

@app.route('/Upload_Files_page',methods=['GET','POST'])
def Upload_Files_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/Locations.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""

@app.route('/Generate_Encodings_page',methods=['GET','POST'])
def Generate_Encodings_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/Generate_Encodes.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""

@app.route('/Delete_Enco_page',methods=['GET','POST'])
def Delete_Enco_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/Delete_Encodes.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/Seperate_Encodings_page',methods=['GET','POST'])
def Seperate_Encodings_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            new_dic=Enco_lst()
            new_dic=dict(new_dic)
            for k, v in new_dic.items():
                num = v
                e=k+"\t \t"+ str(num)+"\n"
                if e:
                    flash(e)
                else:
                    flash("No encodings are available.")
            return render_template('/admin/Seperate_Encodes.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/registration_control_page',methods=['GET','POST'])
def Registration_Control_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/Registration_control.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""

@app.route('/view_attendance_page',methods=['GET','POST'])
def view_attendance_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/view_attendance.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route('/Make_Folders',methods=['GET','POST'])
def Make_Folders():
    alert="alert"
    flag=0
    if request.method == "POST":
        dep=request.form["department"].upper()
        year=request.form["year"]
        subjects=request.form["subject_codes"].upper()
        subject_codes=subjects.split(",")
        try:
            hours=mongo.db.hours
            hours_value=hours.find({})
            for hour in hours_value:
                hr=hour["hours"]
        except:
            flag=0
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            dep_path=default+dep+"\\"                
            year_path=dep_path+str(year)+"\\"        
            original_path=year_path+"original\\"     
            encode_err_path=year_path+"error\\"      
            subjects_path=year_path+"subjects\\"     

            if not os.path.isdir(str(dep_path)):    
                os.mkdir(str(dep_path))             
                flag=1
            if not os.path.isdir(str(year_path)):
                os.mkdir(str(year_path))
                flag=1
            if not os.path.isdir(str(original_path)):
                os.mkdir(str(original_path))
                flag=1
            if not os.path.isdir(str(encode_err_path)):
                os.mkdir(str(encode_err_path))
                flag=1
            if not os.path.isdir(str(subjects_path)):
                os.mkdir(str(subjects_path))
                flag=1
            for i in range(len(subject_codes)): 
                if len(subject_codes[i])!=0:            
                    subject_code_path=str(subjects_path+subject_codes[i]+"\\") 
                    if not os.path.isdir(str(subject_code_path)):  
                        os.mkdir(subject_code_path)         
                        flag=1
                    if not os.path.isdir(str(subject_code_path+"error\\")): 
                        os.mkdir(str(subject_code_path+"error\\"))  
                        flag=1
                    if not os.path.isdir(str(subject_code_path+"extracted\\")): 
                        os.mkdir(str(subject_code_path+"extracted\\")) 
                        flag=1
                    if not os.path.isdir(str(subject_code_path+"excel\\")):  
                        os.mkdir(str(subject_code_path+"excel\\")) 
                        flag=1
                    if not os.path.isdir(str(subject_code_path+"group\\")):  
                        os.mkdir(str(subject_code_path+"group\\"))  
                        flag=1
                    try:
                        for j in range(int(hr)):             
                            if not os.path.isdir(str(subject_code_path+"group\\"+str(j+1))):
                                os.mkdir(str(subject_code_path+"group\\"+str(j+1)))          
                                flag=1
                    except:
                        flag=0
                        
            if flag==1:
                flash("Requested folders are ready to use")
                return redirect("/admin_page")
            elif flag==0:
                flash("First make hours ")
                return redirect("/admin_page")
            else:
                flash("Requested folders already exist for "+str(dep)+"-"+year+" year")
                return render_template('/admin/admin.html',username=name,fail=alert)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/admin_page">here</a> to Go back</h2></center>"""


@app.route('/Save_Hour',methods=['GET','POST'])
def Save_Hour():
    alert="alert"
    if 'username' in session:
        users = mongo.db.users
        hours=mongo.db.hours
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            if request.method == "POST":
                new_hour=request.form["hour"]
                hour_value = list(hours.find({}))
                if len(hour_value)!= 0:
                    hours_value = hours.find({})
                    for hour in hours_value:
                        existing_hour = hour["hours"]
                        hours.update_one({"hours":existing_hour},{"$set":{"hours":new_hour}})
                        break
                else:
                    hours.insert_one({"hours" : new_hour})
                flash("Number of hours per day have been set to "+str(new_hour)+" successfully")
                return render_template('/admin/hours.html',username=name)
            else:
                return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Hours_page">here</a> to Go back</h2></center>"""
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Hours_page">here</a> to Go back</h2></center>"""

@app.route('/Upload_Photos',methods=['GET','POST'])
def Upload_Photos():
    win = Tk()
    if request.method == "POST":
        dep=request.form["department"]
        year=request.form["year"]
        path = os.getcwd()+"\\static\\images\\"+dep+"\\"+year+"\\"+"original"
        file = askopenfilenames(parent=win, title='Choose Photos')
        for w in file:
            src_path = str(w)
            dst_path = str(path)
            try:
                shutil.copy(src_path, dst_path)
            except:
                flash("No Folders are created, unable to upload photos ")
        try:    
            win.destroy()
        except:
            print("Window already destroyed")
        flash("Photos upload successful")
        return redirect(url_for("Generate_Encodings_page") )
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Generate_Encodings_page">here</a> to Go back</h2></center>"""

@app.route('/Upload_Template',methods=['GET','POST'])
def Upload_Template():
    win = Tk()
    if request.method == "POST":
        dep=request.form["department"]
        year=request.form["year"]
        subject=request.form["subject_codes"]
        path = os.getcwd()+"\\static\\images\\"+dep+"\\"+year+"\\subjects\\"+subject+"\\excel"
        file = askopenfilenames(parent=win, title='Choose Template')
        for w in file:
            src_path = str(w)
            dst_path = str(path)
            try:
                shutil.copy(src_path, dst_path)
                flash("Template upload successful")
            except Exception as e:
                flash("No Folders are created, unable to upload template ")
        try:    
            win.destroy()
        except:
            print("Window already destroyed")
        
        return redirect(url_for("Upload_Files_page") )
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Generate_Encodings_page">here</a> to Go back</h2></center>"""

@app.route('/Generate_Encodings',methods=['GET','POST'])
def Generate_Encodings():
    alert="alert"
    if request.method == "POST":
        dep=request.form["department"]
        year=request.form["year"]
        opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath=createpath(dep,year,"","")
        if os.path.isdir(opath) and os.path.isdir(erpath):
            for filename in os.listdir(erpath):
                ipath=str(erpath+filename)
                os.remove(ipath)
            error_image_filename=Encodings(opath,erpath,ojsonpath)
            if len(error_image_filename)!=0:
                if 'username' in session:
                    users = mongo.db.users
                    login_user = users.find_one({'username' : session['username']})
                    name=login_user["firstname"]+" "+login_user["lastname"]
                    if login_user["username"] == "ADMIN":
                        flash("Please retake original images for the following students of "+dep+"-"+year)
                        return render_template('/admin/Generate_Encodes.html',username=name,filenames=error_image_filename,fail=alert)
                    else:
                        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/admin_page">here</a> to Go back</h2></center>"""
            else:
                flash("Encodings Generated successfully for the students of "+dep+"-"+year)
                return redirect(url_for("Generate_Encodings_page") )
        else:
            flash("Department "+dep+" year "+year+" folders not created.")
            if 'username' in session:
                users = mongo.db.users
                login_user = users.find_one({'username' : session['username']})
                name=login_user["firstname"]+" "+login_user["lastname"]
                if login_user["username"] == "ADMIN":
                    return render_template('/admin/Generate_Encodes.html',username=name,fail=alert)
                else:
                    return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Generate_Encodings_page">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Generate_Encodings_page">here</a> to Go back</h2></center>"""


@app.route('/Seperate_Encodings',methods=['GET','POST'])
def Seperate_Encodings():
    users = mongo.db.users
    login_user = users.find_one({'username' : session['username']})
    name=login_user["firstname"]+" "+login_user["lastname"]
    if login_user["username"] == "ADMIN":
        if request.method == "POST":
            dep=request.form["department"].upper()
            year=request.form["year"].upper()   
            sub=request.form["subject_code"].upper()
            dep1=request.form["department1"].upper()
            year1=request.form["year1"].upper()   
            sub1=request.form["subject_code"].upper()
            path=default+dep+"\\"+year+"\\subjects\\"+sub+"\\"
            try:
                src_path = os.getcwd()+"\\static\\images\\"+dep+"\\"+year+"\\original.json"
                src_path1 = os.getcwd()+"\\static\\images\\"+dep1+"\\"+year1+"\\original.json"
                dst_path = os.getcwd()+"\\static\\images\\Subject\\"+sub
                if not os.path.exists(dst_path):
                    os.makedirs(dst_path)
            except:
                flash("No Folders are created, unable to separate encodings")
            try:
                files=[src_path,src_path1]
                with open(src_path) as f:
                    jsonstr = json.load(f) 
                with open(src_path1) as f:
                    jsonstr1 = json.load(f) 
                jsonstr.update(jsonstr1)

                with open(dst_path+'\\original.json', 'w') as output_file:
                    json.dump(jsonstr, output_file)
            except:
                flash("No Encodings are generated, first generate encodings")
                flash("First Generate Encodings for individual Dept. and years")
                return render_template('/admin/Seperate_Encodes.html',username=name)
            if os.path.isdir(dst_path):
                flash("Encodings Combined for "+dep+"-"+year+" and "+dep1+"-"+year1)
                flash("Upload Template For Combined Subject"+sub)
                return render_template('/admin/Seperate_Encodes.html',username=name)
            else:
                flash("Invalid details")
            return render_template('/admin/Seperate_Encodes.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Seperate_Encodings_page">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Seperate_Encodings_page">here</a> to Go back</h2></center>"""

def file_copy(source):
    target=project+"templates/"+"register.html"
    shutil.copy(source,target)


@app.route('/Open_registrations',methods=['GET','POST'])
def Open_registrations():
    if request.method == "POST":
        source=project+"templates/"+"register_template.html"
        file_copy(source)
        flash("Registrations are open now")
        return redirect(url_for('Registration_Control_page'))
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Registration_Control_page">here</a> to Go back</h2></center>"""

@app.route('/Close_registrations',methods=['GET','POST'])
def Close_registrations():
    users = mongo.db.users
    OTPS = mongo.db.OTPS
    Admin = users.find_one({'username' : 'ADMIN'})
    if(Admin):
        Admin_mailid = Admin['mailid']
        passcode = OTPS.find_one({'mailid': Admin_mailid})
    if request.method == "POST":
        source=project+"templates/"+"reg_closed_template.html"
        file_copy(source)
        # OTPS.remove_one({'mailid': Admin_mailid },True)
        flash("Registrations are closed successfully")
        return redirect(url_for('Registration_Control_page'))
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Registration_Control_page">here</a> to Go back</h2></center>"""


@app.route('/Generate_Passcode',methods=['GET','POST'])
def Generate_Passcode():
    if request.method == "POST":
        users = mongo.db.users
        OTPS = mongo.db.OTPS
        Admin_valid = users.find_one({'username' : 'ADMIN' })
        Admin_mailid = Admin_valid['mailid']
        existing_mail = OTPS.find_one({'mailid' : Admin_mailid})
        
        if Admin_valid:
            OTPO = mail(Admin_mailid,0)
            OTP = bcrypt.hashpw(OTPO.encode('utf-8'), bcrypt.gensalt())
            if existing_mail:
                OTPS.update_one({'mailid' : Admin_mailid},{"$set" : {"OTP": OTP}})
            else:
                OTPS.insert_one({'OTP' : OTP , 'mailid' : Admin_mailid})
        flash('New passcode to Register : '+OTPO)
        return redirect(url_for('Registration_Control_page'))
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Generate_Passcode">here</a> to Go back</h2></center>"""

@app.route('/Locations_page',methods=['GET','POST'])
def Locations_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/Locations.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Locations_page">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Locations_page">here</a> to Go back</h2></center>"""

@app.route('/Combine_page',methods=['GET','POST'])
def Combine_page():
    if 'username' in session:
        users = mongo.db.users
        login_user = users.find_one({'username' : session['username']})
        name=login_user["firstname"]+" "+login_user["lastname"]
        if login_user["username"] == "ADMIN":
            return render_template('/admin/combine.html',username=name)
        else:
            return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Locations_page">here</a> to Go back</h2></center>"""
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Locations_page">here</a> to Go back</h2></center>"""

@app.route('/Combine_upload',methods=['GET','POST'])
def Combine_upload():
    win = Tk()
    if request.method == "POST":
        subject=request.form["subject"]
        path = os.getcwd()+"\\static\\images\\Subject\\"+subject+"\\excel"
        path1 = os.getcwd()+"\\static\\images\\Subject\\"+subject+"\\error"
        path2 = os.getcwd()+"\\static\\images\\Subject\\"+subject+"\\extracted"
        path3 = os.getcwd()+"\\static\\images\\Subject\\"+subject+"\\group"
        if not os.path.exists(path):
            os.makedirs(path)
        if not os.path.exists(path1):
            os.makedirs(path1)
        if not os.path.exists(path2):
            os.makedirs(path2)
        if not os.path.exists(path3):
            os.makedirs(path3)
        
        file = askopenfilenames(parent=win, title='Choose Template')
        for w in file:
            src_path = str(w)
            dst_path = str(path)
            try:
                shutil.copy(src_path, dst_path)
            except:
                flash("No Folders are created, unable to upload template ")
        try:    
            win.destroy()
        except:
            print("Window already destroyed")
        flash("Template upload successful")
        return redirect(url_for("Combine_page") )
    else:
        return """<center><h2>Unautnorised Access found... You are not an admin<br><br> Click <a href="/Generate_Encodings_page">here</a> to Go back</h2></center>"""


def match_combine(epath,erpath,sub,hour):
    absent=[]
    present=0
    try:
        now = datetime.now()
        time=str(now.strftime("%I:%M %p"))
        project=str(os.getcwd()+"\\")
        default=str(project+"static\\images\\Subject")
        originallist={}
        # for department in departments:
        jsonpath=str(default+sub+"\\original.json")
        data = open(jsonpath,"r")
        n=data.read()
        temp_list=json.loads(n)
        originallist.update(temp_list)
        for i in originallist:
            originallist[i]=numpy.asarray(originallist[i])
        l=[]
        for a in range(0,2):
            #extracted images encoding
            extractedlist=[]
            for filename in os.listdir(epath):
                sublist=[]
                filename=str(filename)
                sublist.append(filename)
                file_path=epath+filename
                image=face_recognition.load_image_file(file_path)
                image_encoding=face_recognition.face_encodings(image,num_jitters=5)
                if len(image_encoding)>0:
                    image_encoding=image_encoding[0]
                    sublist.append(image_encoding)
                    extractedlist.append(sublist)
                else:
                    shutil.move(epath+filename,erpath)
            
            for key in originallist:
                rollno=key[0:10]
                if rollno in l:
                    continue 
                for esublist in extractedlist:
                    oimg=originallist[key]
                    eimg=esublist[1]
                    result=face_recognition.compare_faces([oimg],eimg,tolerance=0.435)
                    if result[0]== True:
                        l.append(rollno)
                        os.remove(epath+esublist[0])
                        extractedlist.remove(esublist)
                        break
        absent=[]
        present=0
        # for department in departments:
        xl_fname=+sub+".html"
        excelpath=str(default+"\\Subject\\"+sub+"\\excel\\")
        xlhtmlpath=str(project+"templates/excel_html/"+xl_fname)
        c,a=create_sheet(l,sub,excelpath,xlhtmlpath,time) 
        absent+=a
        present+=c
        return present,absent
    except:
        flash("Please Separate Encodings first. If you are viewing this message you either haven't \n separated Encodings, if they are separated retake photo again.")
        return present,absent


@app.route('/mark_attendance_combine',methods=['GET','POST'])
def mark_attendance_combine():
    success="success"
    if request.method == "POST":
        subject=request.form["subject"].upper()
        hr=request.form["hour"]
        # opath,ojsonpath,erpath,gpath,epath,extracted_erpath,cjsonpath=createpath(departments[0],year,sub,hr)
        epath=os.getcwd()+"\\static\\images\\Subject\\"+subject+"\\extracted"
        extracted_erpath=os.getcwd()+"\\static\\images\\Subject\\"+subject+"\\error"

        present,l=match_combine(epath,extracted_erpath,subject,hr)
        present=str(present)
        images=[]
        for filename in os.listdir(epath):
            ipath=str("\\static\\images\\Subject\\"+subject+"\\extracted\\"+filename)
            images.append(ipath)
        for filename in os.listdir(extracted_erpath):
            ipath=str("\\static\\images\\Subject\\"+subject+"\\error\\"+filename)
            images.append(ipath)
        flash("Period-"+hr+" Attendance of "+" "  +subject+" subject is marked successfully for "+present+" students")
        return render_template("error.html",dep="dep",departments_list="departments",year="year",sub=subject,hour=hr,present=present,success=success,images=images,absent_list=l)
    else:
        return """<center><h2>Unautnorised Access found...<br><br> Click <a href="/">here</a> to Go back</h2></center>"""


@app.route("/Delete_Enco",methods=['GET','POST'])
def Delete_Enco():
    dep=request.form["department"]
    year=request.form["year"]
    section=request.form["section"]
    path=str(os.getcwd())
    path1=str(os.getcwd())
    try:
        path=path+"\\static\\images\\Subject\\"+section+"\\original.json"
        os.remove(path)
        flash("Encodings Deleted successfully for the students of Subject "+section)
        return redirect(url_for("Delete_Enco_page") )
    except:
        path1=path1+"\\static\\images\\"+dep+"\\"+year+"\\original.json"
        try:
            os.remove(path1)
        except:
            flash("No Encodings exist for given details")
            return redirect(url_for("Delete_Enco_page") )
        flash("Encodings Deleted successfully for the students of "+dep+"-"+year)
        return redirect(url_for("Delete_Enco_page") )


if __name__ == "__main__":
    app.run(host=ip_addr,port="5000",use_reloader=True,debug=True)

